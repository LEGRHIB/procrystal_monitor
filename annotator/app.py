#!/usr/bin/env python3
"""
crystal_annotator/app.py
────────────────────────
Minimal Flask web app for manual crystallization annotation.

Run:
    python app.py
    open http://localhost:8000
    ~/OneDrive - KU Leuven/DATA/experiments/<ID>/raw_images/img_YYYYMMDD_HHMMSS.png
"""

import csv
import io
import json
import math
import os
import re
import signal
import time
import threading
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import cv2
    import numpy as np
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

try:
    import serial
    import serial.tools.list_ports
    HAS_SERIAL = True
except ImportError:
    HAS_SERIAL = False

import glob as globmod  # renamed to avoid clash with flask

from flask import Flask, jsonify, render_template, request, send_file, abort, Response

# ─── configuration ───────────────────────────────────────────────────────────

_default_root = Path.home() / "OneDrive - KU Leuven" / "DATA" / "experiments"
DATA_ROOT = Path(os.environ.get("CRYSTAL_DATA_ROOT", str(_default_root)))
# Match img_YYYYMMDD_HHMMSS.{png,jpg,jpeg} (strict formatting)
FNAME_RE = re.compile(r"^img_(\d{8})_(\d{6})\.(png|jpg|jpeg)$", re.IGNORECASE)
# Accept any image extension
IMAGE_EXTS = {".png", ".jpg", ".jpeg"}

app = Flask(__name__)


# ─── droplet detection ───────────────────────────────────────────────────────

def detect_droplet_boundaries(image):
    """Find top and bottom meniscus lines of the droplet using intensity profile."""
    if not HAS_CV2:
        return None, None
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image.copy()
    h = gray.shape[0]
    row_profile = np.mean(gray, axis=1)
    kernel = np.ones(11) / 11
    smoothed = np.convolve(row_profile, kernel, mode='same')
    bright_rows = np.where(smoothed > np.mean(smoothed))[0]
    if len(bright_rows) < 2:
        margin = int(h * 0.15)
        return margin, h - margin
    top = max(0, int(bright_rows[0]) - 10)
    bottom = min(h, int(bright_rows[-1]) + 10)
    return top, bottom


def apply_enhancement(image):
    """
    Denoise → background-subtract → percentile-stretch pipeline.
    Works better than CLAHE for low-SNR tube images with uneven illumination:
      1. bilateralFilter  — remove grain while keeping crystal/meniscus edges
      2. large GaussianBlur → background estimate (slow illumination gradient)
      3. subtract background + 128 offset → flat mid-gray field, crystals visible
      4. percentile stretch (2–98) — use full dynamic range without noise boost
    """
    # 1. Denoise — edge-preserving
    denoised = cv2.bilateralFilter(image, d=9, sigmaColor=50, sigmaSpace=50)

    # 2. Background estimate via very large blur
    bg = cv2.GaussianBlur(denoised.astype(np.float32), (0, 0), sigmaX=40)

    # 3. Background subtraction centred on mid-gray
    corrected = denoised.astype(np.float32) - bg + 128.0
    corrected  = np.clip(corrected, 0, 255)

    # 4. Gentle contrast stretch
    p2, p98 = np.percentile(corrected, 2), np.percentile(corrected, 98)
    if p98 > p2:
        corrected = (corrected - p2) / (p98 - p2) * 255.0
        corrected = np.clip(corrected, 0, 255)

    return corrected.astype(np.uint8)


# In-memory cache: {(exp_name, ref_filename): np.ndarray}
_ref_cache: dict = {}

# Export progress: {exp_name: {current, total, done}}
_export_progress: dict = {}


def align_to_reference(ref_gray_f32: "np.ndarray", target_img: "np.ndarray",
                       drift_clamp_px: float = 50.0) -> "np.ndarray":
    """Phase-correlation drift correction (translation only, clamped to ±drift_clamp_px).
    Returns target_img warped so it aligns with ref_gray_f32."""
    tgt_gray = cv2.cvtColor(target_img, cv2.COLOR_BGR2GRAY).astype(np.float32)
    (dx, dy), _ = cv2.phaseCorrelate(ref_gray_f32, tgt_gray)
    dx = max(-drift_clamp_px, min(drift_clamp_px, dx))
    dy = max(-drift_clamp_px, min(drift_clamp_px, dy))
    M = np.float32([[1, 0, -dx], [0, 1, -dy]])
    h, w = target_img.shape[:2]
    return cv2.warpAffine(target_img, M, (w, h),
                          flags=cv2.INTER_LINEAR,
                          borderMode=cv2.BORDER_REFLECT_101)


def apply_temporal_enhancement(ref_img: "np.ndarray", current_img: "np.ndarray") -> "np.ndarray":
    """Align current frame to reference, subtract background, percentile-stretch.
    Highlights changes (crystals growing) against the static tube background."""
    ref_gray_f32 = cv2.cvtColor(ref_img, cv2.COLOR_BGR2GRAY).astype(np.float32)
    aligned = align_to_reference(ref_gray_f32, current_img)
    diff = aligned.astype(np.float32) - ref_img.astype(np.float32) + 128.0
    diff = np.clip(diff, 0, 255)
    p1, p99 = np.percentile(diff, 1), np.percentile(diff, 99)
    if p99 > p1:
        diff = (diff - p1) / (p99 - p1) * 255.0
        diff = np.clip(diff, 0, 255)
    return diff.astype(np.uint8)


def _incremental_export_frame(exp_name: str, frame_idx: int):
    """Update dataset/coco_crystals.json for a single frame after annotation changes.
    Only runs if dataset/ folder already exists (full export was done first).
    Thread-safe via a per-experiment lock."""
    if not HAS_CV2:
        return
    dataset_dir = DATA_ROOT / exp_name / "dataset"
    coco_path = dataset_dir / "coco_crystals.json"
    if not dataset_dir.exists() or not coco_path.exists():
        return  # no dataset yet — user must do full export first

    try:
        frames = scan_experiment(DATA_ROOT / exp_name)
        cal = get_tube_width_calibration(exp_name)
        pixels_per_mm = cal.get("pixels_per_mm")
        lpath = annotations_dir(exp_name) / "lengths.jsonl"

        # Load existing COCO
        coco = json.loads(coco_path.read_text())

        # Remove all annotations for this frame
        coco["annotations"] = [
            a for a in coco.get("annotations", [])
            if a.get("image_id") != frame_idx
        ]

        # Add fresh annotations for this frame from lengths.jsonl
        if lpath.exists():
            next_id = max((a["id"] for a in coco["annotations"]), default=0) + 1
            for line in lpath.read_text().splitlines():
                line = line.strip()
                if not line:
                    continue
                m = json.loads(line)
                if m.get("frame_idx") != frame_idx:
                    continue
                cid = m.get("crystal_id", "")
                if cid.startswith("__"):
                    continue
                p1, p2 = m.get("p1", {}), m.get("p2", {})
                if not p1 or not p2:
                    continue
                x_min = min(p1["x"], p2["x"])
                y_min = min(p1["y"], p2["y"])
                bw = abs(p2["x"] - p1["x"])
                bh = abs(p2["y"] - p1["y"])
                ann = {
                    "id": next_id,
                    "image_id": frame_idx,
                    "category_id": 1,
                    "crystal_id": cid,
                    "direction": m.get("direction"),
                    "bbox": [x_min, y_min, bw, bh],
                    "h_px": m.get("h_px"),
                    "v_px": m.get("v_px"),
                    "measurement_id": m.get("id"),
                }
                if pixels_per_mm:
                    if m.get("h_px"):
                        ann["h_um"] = round(m["h_px"] / pixels_per_mm * 1000, 2)
                    if m.get("v_px"):
                        ann["v_um"] = round(m["v_px"] / pixels_per_mm * 1000, 2)
                coco["annotations"].append(ann)
                next_id += 1

        coco_path.write_text(json.dumps(coco, indent=2))

        # Re-save crop image for this frame if crop_region is stored in settings
        settings = load_settings(exp_name)
        crop_region = settings.get("crop_region")
        if crop_region and frame_idx < len(frames):
            frame_file = frames[frame_idx]["filename"]
            img = cv2.imread(str(DATA_ROOT / exp_name / "raw_images" / frame_file))
            if img is not None:
                cx = int(crop_region.get("x", 0))
                cy = int(crop_region.get("y", 0))
                cw = int(crop_region.get("w", img.shape[1]))
                ch = int(crop_region.get("h", img.shape[0]))
                crop = img[cy:cy + ch, cx:cx + cw]
                stem = frame_file.replace(".png", "")
                crops_dir = dataset_dir / "crops"
                if crops_dir.exists():
                    cv2.imwrite(str(crops_dir / f"{stem}.png"), crop)
    except Exception:
        pass  # best-effort, don't disrupt the annotation workflow


def _build_coco_crystals(exp_name: str, dataset_dir: Path, frames: list, cal: dict):
    """Build coco_crystals.json from lengths.jsonl measurements."""
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return
    pixels_per_mm = cal.get("pixels_per_mm")
    coco = {
        "info": {"description": f"Crystal measurements — {exp_name}", "version": "1.0"},
        "categories": [{"id": 1, "name": "crystal", "supercategory": "none"}],
        "images": [{"id": f["frame_idx"], "file_name": f["filename"]} for f in frames],
        "annotations": [],
    }
    ann_id = 1
    for line in lpath.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        m = json.loads(line)
        cid = m.get("crystal_id", "")
        if cid.startswith("__"):
            continue
        p1, p2 = m.get("p1", {}), m.get("p2", {})
        if not p1 or not p2:
            continue
        x_min = min(p1["x"], p2["x"])
        y_min = min(p1["y"], p2["y"])
        bw = abs(p2["x"] - p1["x"])
        bh = abs(p2["y"] - p1["y"])
        ann = {
            "id": ann_id,
            "image_id": m.get("frame_idx"),
            "category_id": 1,
            "crystal_id": cid,
            "direction": m.get("direction"),
            "bbox": [x_min, y_min, bw, bh],
            "h_px": m.get("h_px"),
            "v_px": m.get("v_px"),
        }
        if pixels_per_mm:
            if m.get("h_px"):
                ann["h_um"] = round(m["h_px"] / pixels_per_mm * 1000, 2)
            if m.get("v_px"):
                ann["v_um"] = round(m["v_px"] / pixels_per_mm * 1000, 2)
        coco["annotations"].append(ann)
        ann_id += 1
    (dataset_dir / "coco_crystals.json").write_text(json.dumps(coco, indent=2))


# ─── helpers ─────────────────────────────────────────────────────────────────

def parse_timestamp(fname: str):
    """Return datetime from 'img_YYYYMMDD_HHMMSS.{png,jpg,jpeg}', or None if unparseable."""
    m = FNAME_RE.match(fname)
    if m is None:
        return None
    try:
        return datetime.strptime(f"{m.group(1)}_{m.group(2)}", "%Y%m%d_%H%M%S")
    except ValueError:
        return None


def scan_experiment(exp_path: Path) -> list[dict]:
    """Return sorted list of {filename, timestamp_iso, t_sec, frame_idx}.
    Accepts any image file (png, jpg, jpeg) with or without img_YYYYMMDD_HHMMSS naming.
    For files without parseable timestamps, uses file modification time as fallback.
    """
    raw = exp_path / "raw_images"
    if not raw.is_dir():
        return []

    frames = []
    for f in sorted(raw.iterdir()):
        # Accept any image extension
        if f.suffix.lower() not in IMAGE_EXTS:
            continue
        
        # Try to parse timestamp from filename
        ts = parse_timestamp(f.name)
        
        # Fallback to file modification time if timestamp parse fails
        if ts is None:
            try:
                ts = datetime.fromtimestamp(f.stat().st_mtime)
            except (OSError, ValueError):
                # If we can't get mtime, skip this file
                continue
        
        frames.append({"filename": f.name, "timestamp": ts})

    frames.sort(key=lambda r: r["timestamp"])
    if not frames:
        return []

    t0 = frames[0]["timestamp"]
    result = []
    for i, f in enumerate(frames):
        dt = (f["timestamp"] - t0).total_seconds()
        result.append({
            "frame_idx": i,
            "filename": f["filename"],
            "timestamp_iso": f["timestamp"].isoformat(),
            "t_sec": dt,
            "t_min": round(dt / 60.0, 4),
        })
    return result


def list_experiments() -> list[str]:
    """Return folder names that contain a raw_images/ subdirectory.
    Supports both flat (exp/raw_images/) and nested (exp/droplet_001/raw_images/)."""
    if not DATA_ROOT.is_dir():
        return []
    exps = []
    for d in sorted(DATA_ROOT.iterdir()):
        if not d.is_dir():
            continue
        if (d / "raw_images").is_dir():
            exps.append(d.name)
        # Also check one level deeper for grouped experiments (gantry droplets)
        for sub in sorted(d.iterdir()):
            if sub.is_dir() and (sub / "raw_images").is_dir():
                exps.append(f"{d.name}/{sub.name}")
    return exps


def annotations_dir(exp_name: str) -> Path:
    p = DATA_ROOT / exp_name / "annotations"
    p.mkdir(parents=True, exist_ok=True)
    return p


def results_dir(exp_name: str) -> Path:
    p = DATA_ROOT / exp_name / "results"
    p.mkdir(parents=True, exist_ok=True)
    return p


_SETTINGS_DEFAULTS = {
    "tube_width_mm": 1.5,
    "droplet_diameter_mm": 1.0,
    "drift_clamp_px": 50,
    "crop_region": None,
}


def load_settings(exp_name: str) -> dict:
    spath = DATA_ROOT / exp_name / "settings.json"
    if spath.exists():
        try:
            saved = json.loads(spath.read_text())
            return {**_SETTINGS_DEFAULTS, **saved}
        except Exception:
            pass
    return dict(_SETTINGS_DEFAULTS)


def save_settings(exp_name: str, data: dict):
    spath = DATA_ROOT / exp_name / "settings.json"
    current = load_settings(exp_name)
    current.update(data)
    spath.write_text(json.dumps(current, indent=2))


def _migrate_lengths_ids(exp_name: str):
    """Back-fill UUID ids into any length records that predate the UUID system.
    Idempotent — safe to call on every read/delete."""
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return
    raw = lpath.read_text().splitlines()
    updated = []
    changed = False
    for line in raw:
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        if "id" not in obj:
            obj["id"] = str(uuid.uuid4())
            changed = True
        updated.append(json.dumps(obj))
    if changed:
        lpath.write_text("\n".join(updated) + "\n")


def get_tube_width_calibration(exp_name: str) -> dict:
    """Extract calibration from __TUBE_WIDTH__ measurement in lengths.jsonl.
    Tube width is read from settings (default 1.5 mm).
    Returns {pixels_per_mm: float} or empty dict if not calibrated."""
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return {}

    settings = load_settings(exp_name)
    tube_width_mm = settings.get("tube_width_mm", 1.5)

    for line in lpath.read_text().splitlines():
        line = line.strip()
        if line:
            obj = json.loads(line)
            if obj.get("crystal_id") == "__TUBE_WIDTH__":
                tube_width_px = obj.get("h_px") or obj.get("v_px")
                if tube_width_px and tube_width_px > 0 and tube_width_mm > 0:
                    pixels_per_mm = tube_width_px / tube_width_mm
                    p1 = obj.get("p1", {})
                    p2 = obj.get("p2", {})
                    result = {
                        "pixels_per_mm": pixels_per_mm,
                        "tube_width_px": tube_width_px,
                        "tube_width_mm": tube_width_mm,
                    }
                    if p1 and p2:
                        result["tube_x1"] = int(min(p1.get("x", 0), p2.get("x", 0)))
                        result["tube_x2"] = int(max(p1.get("x", 0), p2.get("x", 0)))
                    return result
    return {}


def get_droplet_height(exp_name: str) -> dict:
    """
    Extract droplet height from lengths.jsonl.
    Returns {height_px, height_um, volume_ul} or empty dict.
    Droplet diameter is read from settings (default 1.0 mm).
    """
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return {}

    calibration = get_tube_width_calibration(exp_name)
    if not calibration:
        return {}

    pixels_per_mm = calibration["pixels_per_mm"]
    settings = load_settings(exp_name)
    droplet_diameter_mm = settings.get("droplet_diameter_mm", 1.0)
    radius_mm = droplet_diameter_mm / 2.0

    for line in lpath.read_text().splitlines():
        line = line.strip()
        if line:
            obj = json.loads(line)
            if obj.get("crystal_id") == "__DROPLET_HEIGHT__":
                height_px = obj.get("v_px", 0)
                if height_px > 0 and pixels_per_mm > 0:
                    height_mm = height_px / pixels_per_mm
                    height_um = height_mm * 1000
                    # Volume = π * r² * h  (cylindrical approximation), result in µL (mm³)
                    volume_ul = math.pi * (radius_mm ** 2) * height_mm
                    return {
                        "height_px": height_px,
                        "height_mm": round(height_mm, 3),
                        "height_um": round(height_um, 2),
                        "volume_ul": round(volume_ul, 3),
                        "droplet_diameter_mm": droplet_diameter_mm,
                    }

    return {}


def export_metrics_excel(exp_name: str) -> bytes:
    """Generate Excel file with dashboard metrics (matches JS dashboard exactly)."""
    if not HAS_OPENPYXL:
        return None

    frames = scan_experiment(DATA_ROOT / exp_name)
    crystals_data = load_crystals(exp_name)
    crystals = crystals_data.get("crystals", [])
    lpath = annotations_dir(exp_name) / "lengths.jsonl"

    # Load measurements
    all_meas = []
    if lpath.exists():
        for line in lpath.read_text().splitlines():
            if line.strip():
                obj = json.loads(line)
                # Apply µm conversion if calibrated
                calibration = get_tube_width_calibration(exp_name)
                pixels_per_mm = calibration.get("pixels_per_mm")
                if pixels_per_mm:
                    if obj.get("h_px"):
                        obj["h_um"] = round(obj["h_px"] / pixels_per_mm * 1000, 2)
                    if obj.get("v_px"):
                        obj["v_um"] = round(obj["v_px"] / pixels_per_mm * 1000, 2)
                all_meas.append(obj)

    calibration = get_tube_width_calibration(exp_name)
    droplet = get_droplet_height(exp_name)
    pixels_per_mm = calibration.get("pixels_per_mm")
    use_um = bool(pixels_per_mm)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Summary sheet
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "Experiment Summary"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    ws[f"A{row}"] = "Total Time (hr)"
    ws[f"B{row}"] = round((frames[-1]["t_min"] / 60) if frames else 0, 3)

    row += 1
    ws[f"A{row}"] = "Nucleation Events"
    ws[f"B{row}"] = len([c for c in crystals if c.get("nucleation_frame_idx") is not None])

    row += 1
    ws[f"A{row}"] = "Droplet Volume (µL)"
    ws[f"B{row}"] = droplet.get("volume_ul", "—")

    row += 1
    ws[f"A{row}"] = "Nucleation Rate (events/hr/µL)"
    if droplet.get("volume_ul") and frames:
        total_hr = (frames[-1]["t_min"] / 60) if frames else 0
        nuc_count = len([c for c in crystals if c.get("nucleation_frame_idx") is not None])
        nuc_rate = (nuc_count / (total_hr * droplet["volume_ul"])) if total_hr > 0 else 0
        ws[f"B{row}"] = round(nuc_rate, 6)
    else:
        ws[f"B{row}"] = "—"

    row += 1
    ws[f"A{row}"] = "Calibration (px/mm)"
    ws[f"B{row}"] = round(pixels_per_mm, 2) if pixels_per_mm else "—"

    # Metrics sheet
    ws2 = wb.create_sheet("Growth Metrics")

    # Header
    h_unit = "µm" if use_um else "px"
    v_unit = "µm" if use_um else "px"
    area_unit = "µm²" if use_um else "px²"
    rate_unit = "µm/hr" if use_um else "px/hr"

    headers = [
        "Crystal",
        f"H size ({h_unit})",
        f"V size ({v_unit})",
        f"H rate ({rate_unit})",
        f"V rate ({rate_unit})",
        f"Area ({area_unit})",
        "Aspect Ratio",
        "Induction Time (hr)"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Helper functions matching JS
    def linear_slope(points):
        if len(points) < 2:
            return None
        n = len(points)
        sx = sum(p[0] for p in points)
        sy = sum(p[1] for p in points)
        sxx = sum(p[0]**2 for p in points)
        sxy = sum(p[0]*p[1] for p in points)
        denom = n*sxx - sx*sx
        if denom == 0:
            return None
        return (n*sxy - sx*sy) / denom

    def latest_value(meas, key, frame_list):
        """Find latest value by timestamp (not file order)."""
        best = None
        for m in meas:
            if m.get(key) is None:
                continue
            frame_idx = m.get("frame_idx")
            if frame_idx is None or frame_idx >= len(frame_list):
                continue
            t_min = frame_list[frame_idx].get("t_min", 0)
            if best is None or t_min > best["t"]:
                best = {"t": t_min, "v": m[key]}
        return best["v"] if best else None

    # Data for each crystal
    for row_idx, crystal in enumerate(crystals, 2):
        cid = crystal["id"]
        c_meas = [m for m in all_meas if m.get("crystal_id") == cid]

        ws2.cell(row=row_idx, column=1).value = cid

        # Latest H and V (by timestamp, using same logic as JS)
        h_key = "h_um" if use_um else "h_px"
        v_key = "v_um" if use_um else "v_px"

        h_latest = latest_value(c_meas, h_key, frames)
        v_latest = latest_value(c_meas, v_key, frames)

        ws2.cell(row=row_idx, column=2).value = round(h_latest, 2) if h_latest else "—"
        ws2.cell(row=row_idx, column=3).value = round(v_latest, 2) if v_latest else "—"

        # Growth rates (slope) - match JS logic
        h_points = []
        v_points = []
        for m in c_meas:
            frame_idx = m.get("frame_idx")
            if frame_idx is None or frame_idx >= len(frames):
                continue
            t_min = frames[frame_idx].get("t_min", 0)
            t_hr = t_min / 60

            h_val = m.get(h_key)
            v_val = m.get(v_key)

            if h_val is not None:
                h_points.append((t_hr, h_val))
            if v_val is not None:
                v_points.append((t_hr, v_val))

        h_slope = linear_slope(h_points)
        v_slope = linear_slope(v_points)

        ws2.cell(row=row_idx, column=4).value = round(h_slope, 2) if h_slope else "—"
        ws2.cell(row=row_idx, column=5).value = round(v_slope, 2) if v_slope else "—"

        # Area (ellipse) - use math.pi to match JS Math.PI
        if h_latest and v_latest:
            area = math.pi * (h_latest / 2) * (v_latest / 2)
            ws2.cell(row=row_idx, column=6).value = round(area, 2)
        else:
            ws2.cell(row=row_idx, column=6).value = "—"

        # Aspect ratio
        if h_latest and v_latest:
            ws2.cell(row=row_idx, column=7).value = round(h_latest / v_latest, 3)
        else:
            ws2.cell(row=row_idx, column=7).value = "—"

        # Induction time
        if crystal.get("nucleation_frame_idx") is not None:
            frame_idx = crystal["nucleation_frame_idx"]
            if frame_idx < len(frames):
                t_min = frames[frame_idx]["t_min"]
                t_hr = t_min / 60
                ws2.cell(row=row_idx, column=8).value = round(t_hr, 3)
        if crystal.get("nucleation_frame_idx") is None:
            ws2.cell(row=row_idx, column=8).value = "Not marked"

    # Auto-size columns
    for ws_sheet in [ws, ws2]:
        for column in ws_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_sheet.column_dimensions[column_letter].width = max_length + 2

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─── routes: pages ───────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ─── routes: API ─────────────────────────────────────────────────────────────

@app.get("/api/experiments")
def api_experiments():
    return jsonify(list_experiments())


@app.get("/api/frames/<path:exp_name>")
def api_frames(exp_name: str):
    exp_path = DATA_ROOT / exp_name
    if not exp_path.is_dir():
        abort(404)
    frames = scan_experiment(exp_path)
    return jsonify(frames)


@app.get("/api/image/<path:exp_name>/<filename>")
def api_image(exp_name: str, filename: str):
    img_path = DATA_ROOT / exp_name / "raw_images" / filename
    if not img_path.is_file():
        abort(404)
    return send_file(img_path, mimetype="image/png")


@app.get("/api/calibration/<path:exp_name>")
def api_get_calibration(exp_name: str):
    """Get calibration (pixels_per_mm) extracted from __TUBE_WIDTH__ measurement."""
    cal = get_tube_width_calibration(exp_name)
    return jsonify(cal)


@app.get("/api/droplet/<path:exp_name>")
def api_get_droplet(exp_name: str):
    """Get droplet height and volume data."""
    droplet = get_droplet_height(exp_name)
    return jsonify(droplet)


@app.get("/api/detect_droplet/<path:exp_name>/<filename>")
def api_detect_droplet(exp_name: str, filename: str):
    """Auto-detect droplet bounding box. Returns {x, y, w, h} in original image pixels."""
    if not HAS_CV2:
        return jsonify({"error": "opencv-python not installed — run: pip install opencv-python numpy"}), 501

    img_path = DATA_ROOT / exp_name / "raw_images" / filename
    if not img_path.is_file():
        abort(404)

    img = cv2.imread(str(img_path))
    if img is None:
        abort(404)

    img_w = img.shape[1]
    cal = get_tube_width_calibration(exp_name)

    # Use tube x-coords from calibration if available, otherwise centre 60%
    x1 = cal.get("tube_x1", int(img_w * 0.2))
    x2 = cal.get("tube_x2", int(img_w * 0.8))
    x1 = max(0, min(x1, img_w - 1))
    x2 = max(x1 + 1, min(x2, img_w))

    top, bottom = detect_droplet_boundaries(img[:, x1:x2])

    pad_x = max(10, int((x2 - x1) * 0.1))
    return jsonify({
        "x": max(0, x1 - pad_x),
        "y": top,
        "w": min(img_w, x2 + pad_x) - max(0, x1 - pad_x),
        "h": bottom - top,
    })


@app.get("/api/enhanced_crop/<path:exp_name>/<filename>")
def api_enhanced_crop(exp_name: str, filename: str):
    """Return a temporally-enhanced PNG for the given crop region.
    Query params: x, y, w, h (in original image pixels); ref=<filename> (reference frame)."""
    if not HAS_CV2:
        return jsonify({"error": "opencv-python not installed"}), 501

    img_path = DATA_ROOT / exp_name / "raw_images" / filename
    if not img_path.is_file():
        abort(404)

    img = cv2.imread(str(img_path))
    if img is None:
        abort(404)

    img_h_full, img_w_full = img.shape[:2]
    try:
        x = int(request.args.get("x", 0))
        y = int(request.args.get("y", 0))
        w = int(request.args.get("w", img_w_full))
        h = int(request.args.get("h", img_h_full))
    except (ValueError, TypeError):
        x, y, w, h = 0, 0, img_w_full, img_h_full

    x  = max(0, min(x, img_w_full - 1))
    y  = max(0, min(y, img_h_full - 1))
    x2 = max(x + 1, min(x + w, img_w_full))
    y2 = max(y + 1, min(y + h, img_h_full))

    # Load reference frame (default: frame 0 of this experiment)
    ref_filename = request.args.get("ref", None)
    if not ref_filename:
        exp_frames = scan_experiment(DATA_ROOT / exp_name)
        ref_filename = exp_frames[0]["filename"] if exp_frames else filename

    cache_key = (exp_name, ref_filename)
    ref_img = _ref_cache.get(cache_key)
    if ref_img is None:
        ref_path = DATA_ROOT / exp_name / "raw_images" / ref_filename
        ref_img = cv2.imread(str(ref_path))
        if ref_img is not None:
            _ref_cache[cache_key] = ref_img

    crop = img[y:y2, x:x2]
    if ref_img is not None and filename != ref_filename:
        ref_crop = ref_img[y:y2, x:x2]
        enhanced = apply_temporal_enhancement(ref_crop, crop)
    else:
        enhanced = apply_enhancement(crop)  # fallback when viewing the ref frame itself

    ok, buf = cv2.imencode(".png", enhanced)
    if not ok:
        abort(500)
    return send_file(io.BytesIO(buf.tobytes()), mimetype="image/png")


@app.post("/api/export/<path:exp_name>/dataset")
def api_export_dataset(exp_name: str):
    """Batch-export crops, CLAHE-enhanced images, and COCO JSON into <exp>/dataset/."""
    if not HAS_CV2:
        return jsonify({"error": "opencv-python not installed — run: pip install opencv-python numpy"}), 501

    exp_path = DATA_ROOT / exp_name
    if not exp_path.is_dir():
        abort(404)

    frames = scan_experiment(exp_path)
    if not frames:
        return jsonify({"error": "No frames found in experiment"}), 400

    dataset_dir  = exp_path / "dataset"
    crops_dir    = dataset_dir / "crops"
    enhanced_dir = dataset_dir / "enhanced"
    crops_dir.mkdir(parents=True, exist_ok=True)
    enhanced_dir.mkdir(parents=True, exist_ok=True)

    cal = get_tube_width_calibration(exp_name)

    # Load frame 0 as the temporal reference for background subtraction
    ref_img = cv2.imread(str(exp_path / "raw_images" / frames[0]["filename"]))

    # Record first-frame crop region in settings for incremental sync
    settings = load_settings(exp_name)
    # crop_region will be updated below after detecting boundaries for frame 0

    coco_droplets = {
        "info": {"description": f"Droplet bounding boxes — {exp_name}", "version": "1.0"},
        "categories": [{"id": 1, "name": "droplet", "supercategory": "none"}],
        "images": [],
        "annotations": [],
    }
    ann_id    = 1
    processed = 0
    errors    = []
    first_crop_saved = False

    _export_progress[exp_name] = {"current": 0, "total": len(frames), "done": False}

    for frame in frames:
        img_path = exp_path / "raw_images" / frame["filename"]
        img = cv2.imread(str(img_path))
        if img is None:
            errors.append(frame["filename"])
            _export_progress[exp_name]["current"] += 1
            continue

        img_w_px = img.shape[1]

        x1_t = cal.get("tube_x1", int(img_w_px * 0.2))
        x2_t = cal.get("tube_x2", int(img_w_px * 0.8))
        x1_t = max(0, min(x1_t, img_w_px - 1))
        x2_t = max(x1_t + 1, min(x2_t, img_w_px))

        top, bottom = detect_droplet_boundaries(img[:, x1_t:x2_t])
        pad_x = max(10, int((x2_t - x1_t) * 0.1))

        cx = max(0, x1_t - pad_x)
        cy = top
        cw = min(img_w_px, x2_t + pad_x) - cx
        ch = bottom - cy

        # Save the crop region for the first frame into settings for incremental sync
        if not first_crop_saved:
            save_settings(exp_name, {"crop_region": {"x": cx, "y": cy, "w": cw, "h": ch}})
            first_crop_saved = True

        crop = img[cy:cy + ch, cx:cx + cw]
        if ref_img is not None and frame["filename"] != frames[0]["filename"]:
            ref_crop = ref_img[cy:cy + ch, cx:cx + cw]
            enhanced = apply_temporal_enhancement(ref_crop, crop)
        else:
            enhanced = apply_enhancement(crop)

        stem = frame["filename"].replace(".png", "")
        cv2.imwrite(str(crops_dir    / f"{stem}.png"), crop)
        cv2.imwrite(str(enhanced_dir / f"{stem}.png"), enhanced)

        coco_droplets["images"].append({
            "id":             frame["frame_idx"],
            "file_name":      f"{stem}.png",
            "width":          cw,
            "height":         ch,
            "crop_x":         cx,
            "crop_y":         cy,
            "timestamp_iso":  frame["timestamp_iso"],
            "t_min":          frame["t_min"],
        })
        coco_droplets["annotations"].append({
            "id":          ann_id,
            "image_id":    frame["frame_idx"],
            "category_id": 1,
            "bbox":        [0, 0, cw, ch],
            "area":        cw * ch,
            "iscrowd":     0,
        })
        ann_id    += 1
        processed += 1
        _export_progress[exp_name]["current"] += 1

    (dataset_dir / "coco_droplets.json").write_text(json.dumps(coco_droplets, indent=2))
    _build_coco_crystals(exp_name, dataset_dir, frames, cal)
    _export_progress[exp_name]["done"] = True

    return jsonify({
        "ok":         True,
        "processed":  processed,
        "errors":     errors,
        "dataset_dir": str(dataset_dir),
    })


@app.get("/api/export/<path:exp_name>/metrics")
def api_export_metrics(exp_name: str):
    """Export dashboard metrics as Excel file and save a copy to results/."""
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl not installed"}), 500

    exp_path = DATA_ROOT / exp_name
    if not exp_path.is_dir():
        abort(404)

    excel_data = export_metrics_excel(exp_name)
    if not excel_data:
        return jsonify({"error": "Failed to generate Excel"}), 500

    # Auto-save a copy to <experiment>/results/
    fname = f"{exp_name}_metrics.xlsx"
    save_path = results_dir(exp_name) / fname
    save_path.write_bytes(excel_data)

    # Send file from disk
    return send_file(
        str(save_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
    )


# ── crystals ──────────────────────────────────────────────────────────────────

def crystals_file(exp_name: str) -> Path:
    """Return path to crystals.json for an experiment."""
    return annotations_dir(exp_name) / "crystals.json"


def load_crystals(exp_name: str) -> dict:
    """Load crystals data, return {crystals: [...]} or empty structure."""
    cf = crystals_file(exp_name)
    if cf.exists():
        return json.loads(cf.read_text())
    return {"crystals": []}


def save_crystals(exp_name: str, data: dict):
    """Save crystals data."""
    cf = crystals_file(exp_name)
    cf.write_text(json.dumps(data, indent=2))


@app.get("/api/crystals/<path:exp_name>")
def api_get_crystals(exp_name: str):
    """Return list of all crystals in an experiment."""
    data = load_crystals(exp_name)
    return jsonify(data.get("crystals", []))


@app.post("/api/crystals/<path:exp_name>")
def api_create_crystal(exp_name: str):
    """Create a new crystal. Body: {notes: "optional notes"}"""
    data = request.get_json(force=True) or {}
    crystals = load_crystals(exp_name)

    # Generate crystal ID
    existing_ids = [c.get("id") for c in crystals["crystals"]]
    crystal_id = f"crystal_{len(existing_ids) + 1}"

    new_crystal = {
        "id": crystal_id,
        "nucleation_frame_idx": None,
        "nucleation_timestamp_iso": None,
        "notes": data.get("notes", ""),
    }
    crystals["crystals"].append(new_crystal)
    save_crystals(exp_name, crystals)
    return jsonify(new_crystal), 201


@app.delete("/api/crystals/<path:exp_name>/<crystal_id>")
def api_delete_crystal(exp_name: str, crystal_id: str):
    """Delete a crystal and its associated measurements."""
    crystals = load_crystals(exp_name)
    crystals["crystals"] = [c for c in crystals["crystals"] if c["id"] != crystal_id]
    save_crystals(exp_name, crystals)

    # Also delete measurements for this crystal
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if lpath.exists():
        lines = [l for l in lpath.read_text().splitlines() if l.strip()]
        lines = [l for l in lines if json.loads(l).get("crystal_id") != crystal_id]
        lpath.write_text("\n".join(lines) + ("\n" if lines else ""))

    return jsonify({"ok": True})


@app.put("/api/crystals/<path:exp_name>/<crystal_id>/nucleation")
def api_set_crystal_nucleation(exp_name: str, crystal_id: str):
    """Set nucleation time for a crystal. Body: {nucleation_frame_idx: N, nucleation_timestamp_iso: "..."}"""
    data = request.get_json(force=True)
    crystals = load_crystals(exp_name)

    for crystal in crystals["crystals"]:
        if crystal["id"] == crystal_id:
            crystal["nucleation_frame_idx"] = data.get("nucleation_frame_idx")
            crystal["nucleation_timestamp_iso"] = data.get("nucleation_timestamp_iso")
            break

    save_crystals(exp_name, crystals)
    return jsonify({"ok": True})


# ── measurements ──────────────────────────────────────────────────────────────

@app.get("/api/lengths/<path:exp_name>")
def api_get_lengths(exp_name: str):
    """Return all length measurements as a JSON array, optionally filtered by crystal_id.
    Only shows measurements for crystals that are defined in crystals.json.
    Automatically includes µm conversion if tube width is calibrated."""
    crystal_id = request.args.get("crystal_id")
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return jsonify([])

    # Load valid crystal IDs from crystals.json
    crystals = load_crystals(exp_name)
    valid_crystal_ids = {c["id"] for c in crystals.get("crystals", [])}
    # Also allow special measurements
    valid_crystal_ids.add("__TUBE_WIDTH__")
    valid_crystal_ids.add("__DROPLET_HEIGHT__")

    # Get calibration if available
    calibration = get_tube_width_calibration(exp_name)
    pixels_per_mm = calibration.get("pixels_per_mm")

    lines = []
    for line in lpath.read_text().splitlines():
        line = line.strip()
        if line:
            obj = json.loads(line)
            # Skip measurements for crystals that don't exist in crystals.json
            if obj.get("crystal_id") not in valid_crystal_ids:
                continue

            # Add µm conversion if calibrated
            if pixels_per_mm:
                if obj.get("h_px"):
                    obj["h_um"] = round(obj["h_px"] / pixels_per_mm * 1000, 2)
                if obj.get("v_px"):
                    obj["v_um"] = round(obj["v_px"] / pixels_per_mm * 1000, 2)

            if crystal_id is None or obj.get("crystal_id") == crystal_id:
                lines.append(obj)
    return jsonify(lines)


@app.post("/api/lengths/<path:exp_name>")
def api_add_length(exp_name: str):
    """Append one measurement line to lengths.jsonl.
    Body: {crystal_id: "crystal_1", frame_idx: N, ...}
    Returns {ok, id} with the UUID of the saved measurement."""
    data = request.get_json(force=True)
    # Accept client-generated UUID or mint a new one
    data.setdefault("id", str(uuid.uuid4()))
    data.setdefault("created_at", datetime.utcnow().isoformat() + "Z")
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    with lpath.open("a") as fh:
        fh.write(json.dumps(data) + "\n")

    # Incremental sync: update dataset if it already exists (best-effort, non-blocking)
    frame_idx = data.get("frame_idx")
    if frame_idx is not None:
        t = threading.Thread(target=_incremental_export_frame, args=(exp_name, frame_idx), daemon=True)
        t.start()

    return jsonify({"ok": True, "id": data["id"]})


@app.delete("/api/lengths/<path:exp_name>/<crystal_id>/<measurement_id>")
def api_delete_length(exp_name: str, crystal_id: str, measurement_id: str):
    """Delete a single measurement by its UUID id field."""
    _migrate_lengths_ids(exp_name)
    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    if not lpath.exists():
        return jsonify({"error": "No measurements file"}), 404

    all_lines = [l for l in lpath.read_text().splitlines() if l.strip()]
    new_lines = []
    deleted_frame = None
    deleted = False
    for line in all_lines:
        obj = json.loads(line)
        if obj.get("crystal_id") == crystal_id and obj.get("id") == measurement_id:
            deleted = True
            deleted_frame = obj.get("frame_idx")
        else:
            new_lines.append(line)

    if not deleted:
        return jsonify({"error": "Measurement not found"}), 404

    lpath.write_text("\n".join(new_lines) + ("\n" if new_lines else ""))

    # Incremental sync after deletion (best-effort, non-blocking)
    if deleted_frame is not None:
        t = threading.Thread(target=_incremental_export_frame, args=(exp_name, deleted_frame), daemon=True)
        t.start()

    return jsonify({"ok": True})


# ── settings ──────────────────────────────────────────────────────────────────

@app.get("/api/settings/<path:exp_name>")
def api_get_settings(exp_name: str):
    return jsonify(load_settings(exp_name))


@app.post("/api/settings/<path:exp_name>")
def api_save_settings(exp_name: str):
    data = request.get_json(force=True) or {}
    # Validate numeric fields
    for key in ("tube_width_mm", "droplet_diameter_mm", "drift_clamp_px"):
        if key in data:
            try:
                data[key] = float(data[key])
            except (ValueError, TypeError):
                return jsonify({"error": f"Invalid value for {key}"}), 400
    save_settings(exp_name, data)
    return jsonify({"ok": True, "settings": load_settings(exp_name)})


# ── crystal notes PATCH ────────────────────────────────────────────────────────

@app.route("/api/crystals/<path:exp_name>/<crystal_id>", methods=["PATCH"])
def api_patch_crystal(exp_name: str, crystal_id: str):
    """Update notes (or other fields) for a crystal."""
    data = request.get_json(force=True) or {}
    crystals = load_crystals(exp_name)
    found = False
    for crystal in crystals["crystals"]:
        if crystal["id"] == crystal_id:
            if "notes" in data:
                crystal["notes"] = data["notes"]
            found = True
            break
    if not found:
        return jsonify({"error": "Crystal not found"}), 404
    save_crystals(exp_name, crystals)
    return jsonify({"ok": True})


# ── cache reset ────────────────────────────────────────────────────────────────

@app.post("/api/reset_cache/<path:exp_name>")
def api_reset_cache(exp_name: str):
    """Evict all cached reference frames for this experiment."""
    to_remove = [k for k in _ref_cache if k[0] == exp_name]
    for k in to_remove:
        del _ref_cache[k]
    return jsonify({"ok": True, "evicted": len(to_remove)})


# ── export: CSV ────────────────────────────────────────────────────────────────

@app.get("/api/export/<path:exp_name>/csv")
def api_export_csv(exp_name: str):
    """Export all measurements as CSV download."""
    exp_path = DATA_ROOT / exp_name
    if not exp_path.is_dir():
        abort(404)

    frames = scan_experiment(exp_path)
    frame_map = {f["frame_idx"]: f for f in frames}
    cal = get_tube_width_calibration(exp_name)
    pixels_per_mm = cal.get("pixels_per_mm")

    lpath = annotations_dir(exp_name) / "lengths.jsonl"
    rows = []
    if lpath.exists():
        for line in lpath.read_text().splitlines():
            line = line.strip()
            if not line:
                continue
            m = json.loads(line)
            cid = m.get("crystal_id", "")
            if cid.startswith("__"):
                continue
            fi = m.get("frame_idx", "")
            frame = frame_map.get(fi, {})
            h_mm = round(m["h_px"] / pixels_per_mm, 4) if m.get("h_px") and pixels_per_mm else ""
            v_mm = round(m["v_px"] / pixels_per_mm, 4) if m.get("v_px") and pixels_per_mm else ""
            rows.append({
                "id": m.get("id", ""),
                "crystal_id": cid,
                "frame_idx": fi,
                "timestamp_iso": frame.get("timestamp_iso", ""),
                "t_min": frame.get("t_min", ""),
                "direction": m.get("direction", ""),
                "h_px": m.get("h_px", ""),
                "h_mm": h_mm,
                "v_px": m.get("v_px", ""),
                "v_mm": v_mm,
                "created_at": m.get("created_at", ""),
            })

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[
        "id", "crystal_id", "frame_idx", "timestamp_iso", "t_min",
        "direction", "h_px", "h_mm", "v_px", "v_mm", "created_at"
    ])
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)

    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"{exp_name}_measurements.csv",
    )


# ── export: progress ───────────────────────────────────────────────────────────

@app.get("/api/export/<path:exp_name>/progress")
def api_export_progress(exp_name: str):
    """Return current export progress for polling."""
    prog = _export_progress.get(exp_name, {"current": 0, "total": 0, "done": True})
    return jsonify(prog)


# =============================================================================
# GANTRY CONTROLLER — integrated web UI
# =============================================================================

# ── serial port detection ─────────────────────────────────────────────────────

GANTRY_DEFAULTS = {
    "baud": 115200,
    "xy_speed": 3000,
    "settle_time": 0.5,
    "camera_index": 0,
    "warmup_frames": 5,
}

# Global state for the gantry session
_gantry_state: Dict = {
    "ser": None,            # serial.Serial object
    "port": None,
    "current_x": 0.0,
    "current_y": 0.0,
    "connected": False,
    "camera": None,         # cv2.VideoCapture
    "camera_ok": False,
    "positions": [],        # marked droplet positions
    "step": 5.0,            # mm per jog
    "monitor_running": False,
    "monitor_thread": None,
    "monitor_log": [],      # recent log lines
    "led_on": True,         # camera LED state (Dino-Lite)
}
_gantry_lock = threading.Lock()


def _find_printer_port() -> Optional[str]:
    """Auto-detect Ender 5 Pro serial port by probing for Marlin firmware."""
    if not HAS_SERIAL:
        return None
    patterns = ['usbserial', 'wchusbserial', 'SLAB_USB', 'usbmodem']

    # Gather candidate ports (pattern match then glob fallback)
    candidates: list[str] = []
    ports = serial.tools.list_ports.comports()
    for port in ports:
        device = port.device.lower()
        for pat in patterns:
            if pat.lower() in device:
                if port.device not in candidates:
                    candidates.append(port.device)
    for pat in patterns:
        for m in globmod.glob(f"/dev/cu.*{pat}*"):
            if m not in candidates:
                candidates.append(m)

    if not candidates:
        return None

    # Probe each candidate for Marlin firmware (M115)
    baud = GANTRY_DEFAULTS["baud"]
    for dev in candidates:
        try:
            ser = serial.Serial(dev, baud, timeout=3)
            time.sleep(2)  # wait for boot
            # flush startup messages
            while ser.in_waiting:
                ser.readline()
            ser.write(b"M115\n")
            deadline = time.time() + 5
            while time.time() < deadline:
                if ser.in_waiting:
                    line = ser.readline().decode(errors='replace')
                    if 'FIRMWARE' in line.upper() or 'MARLIN' in line.upper():
                        ser.close()
                        return dev
                time.sleep(0.05)
            ser.close()
        except Exception:
            continue

    # Fallback: return the first candidate even without Marlin confirmation
    return candidates[0]


def _gantry_send(cmd: str, timeout: float = 30.0) -> str:
    """Send G-code command and wait for 'ok' from Marlin."""
    ser = _gantry_state["ser"]
    if ser is None or not ser.is_open:
        return ""
    ser.write(f"{cmd}\n".encode())
    start = time.time()
    while time.time() - start < timeout:
        if ser.in_waiting:
            line = ser.readline().decode(errors='replace').strip()
            if line.startswith("ok"):
                return line
            if "Error" in line or "error" in line:
                return line
        time.sleep(0.05)
    return ""


def _gantry_move_to(x: float, y: float):
    """Move gantry to absolute XY."""
    _gantry_send("G90")
    _gantry_send(f"G1 X{x:.2f} Y{y:.2f} F{GANTRY_DEFAULTS['xy_speed']}")
    _gantry_send("M400")
    _gantry_state["current_x"] = x
    _gantry_state["current_y"] = y


def _gantry_move_relative(dx: float = 0, dy: float = 0):
    """Move gantry relative to current position."""
    _gantry_send("G91")
    _gantry_send(f"G1 X{dx:.2f} Y{dy:.2f} F{GANTRY_DEFAULTS['xy_speed']}")
    _gantry_send("M400")
    _gantry_send("G90")
    _gantry_state["current_x"] += dx
    _gantry_state["current_y"] += dy


# ── gantry page ───────────────────────────────────────────────────────────────

@app.get("/gantry")
def gantry_page():
    return render_template("gantry.html")


# ── gantry API routes ─────────────────────────────────────────────────────────

@app.get("/api/gantry/status")
def api_gantry_status():
    """Return current gantry state."""
    return jsonify({
        "connected": _gantry_state["connected"],
        "port": _gantry_state["port"],
        "x": _gantry_state["current_x"],
        "y": _gantry_state["current_y"],
        "step": _gantry_state["step"],
        "camera_ok": _gantry_state["camera_ok"],
        "positions": _gantry_state["positions"],
        "monitor_running": _gantry_state["monitor_running"],
        "monitor_log": _gantry_state["monitor_log"][-50:],
        "led_on": _gantry_state["led_on"],
    })


@app.get("/api/gantry/ports")
def api_gantry_ports():
    """List available serial ports."""
    if not HAS_SERIAL:
        return jsonify({"ports": [], "error": "pyserial not installed"})
    ports = serial.tools.list_ports.comports()
    return jsonify({
        "ports": [{"device": p.device, "description": p.description} for p in ports],
        "auto": _find_printer_port(),
    })


@app.post("/api/gantry/connect")
def api_gantry_connect():
    """Connect to the gantry (Ender 5 Pro) via serial."""
    if not HAS_SERIAL:
        return jsonify({"error": "pyserial not installed. Run: pip install pyserial"}), 501

    data = request.get_json(force=True) or {}
    port = data.get("port", "auto")
    baud = int(data.get("baud", GANTRY_DEFAULTS["baud"]))

    with _gantry_lock:
        # Already connected?
        if _gantry_state["connected"] and _gantry_state["ser"] and _gantry_state["ser"].is_open:
            return jsonify({"ok": True, "port": _gantry_state["port"], "msg": "Already connected"})

        if port == "auto":
            port = _find_printer_port()
            if port is None:
                return jsonify({"error": "Could not auto-detect printer port. Is it plugged in?"}), 400

        try:
            ser = serial.Serial(port, baud, timeout=10)
            time.sleep(2)  # Wait for Marlin boot
            # Flush startup messages
            while ser.in_waiting:
                ser.readline()
            ser.write(b"\n")
            time.sleep(0.5)
            while ser.in_waiting:
                ser.readline()

            _gantry_state["ser"] = ser
            _gantry_state["port"] = port
            _gantry_state["connected"] = True
            _gantry_state["current_x"] = 0.0
            _gantry_state["current_y"] = 0.0

            # Safety: disable heaters
            _gantry_send("M104 S0")
            _gantry_send("M140 S0")

            # Verify Marlin firmware responds
            warning = None
            ser.write(b"M115\n")
            marlin_ok = False
            deadline = time.time() + 5
            while time.time() < deadline:
                if ser.in_waiting:
                    line = ser.readline().decode(errors='replace')
                    if 'FIRMWARE' in line.upper() or 'MARLIN' in line.upper():
                        marlin_ok = True
                        break
                    if line.strip().startswith("ok"):
                        break
                time.sleep(0.05)
            if not marlin_ok:
                warning = (f"Port {port} connected but did not identify as Marlin firmware. "
                           "This may be the wrong device. Try selecting a different port manually.")

            resp = {"ok": True, "port": port}
            if warning:
                resp["warning"] = warning
            return jsonify(resp)
        except Exception as e:
            return jsonify({"error": f"Connection failed: {e}"}), 400


@app.post("/api/gantry/disconnect")
def api_gantry_disconnect():
    """Disconnect from gantry and release camera."""
    with _gantry_lock:
        if _gantry_state["ser"] and _gantry_state["ser"].is_open:
            _gantry_send("M104 S0")
            _gantry_send("M140 S0")
            _gantry_state["ser"].close()
        _gantry_state["ser"] = None
        _gantry_state["connected"] = False
        _gantry_state["port"] = None

        if _gantry_state["camera"] is not None:
            _gantry_state["camera"].release()
            _gantry_state["camera"] = None
            _gantry_state["camera_ok"] = False

    return jsonify({"ok": True})


@app.post("/api/gantry/home")
def api_gantry_home():
    """Home X and Y axes."""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    with _gantry_lock:
        _gantry_send("G28 X Y", timeout=60)
        _gantry_send("G90")
        _gantry_send("M400")
        _gantry_state["current_x"] = 0.0
        _gantry_state["current_y"] = 0.0
    return jsonify({"ok": True, "x": 0.0, "y": 0.0})


@app.post("/api/gantry/jog")
def api_gantry_jog():
    """Jog gantry in a direction. Body: {direction: "x+"|"x-"|"y+"|"y-", step: float}"""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    data = request.get_json(force=True) or {}
    direction = data.get("direction", "")
    step = float(data.get("step", _gantry_state["step"]))
    _gantry_state["step"] = step

    with _gantry_lock:
        if direction == "x+":
            _gantry_move_relative(dx=step)
        elif direction == "x-":
            _gantry_move_relative(dx=-step)
        elif direction == "y+":
            _gantry_move_relative(dy=step)
        elif direction == "y-":
            _gantry_move_relative(dy=-step)
        else:
            return jsonify({"error": f"Invalid direction: {direction}"}), 400

    return jsonify({
        "ok": True,
        "x": round(_gantry_state["current_x"], 2),
        "y": round(_gantry_state["current_y"], 2),
    })


@app.post("/api/gantry/move_to")
def api_gantry_move_to():
    """Move to absolute XY. Body: {x: float, y: float}"""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    data = request.get_json(force=True) or {}
    x = float(data.get("x", _gantry_state["current_x"]))
    y = float(data.get("y", _gantry_state["current_y"]))
    with _gantry_lock:
        _gantry_move_to(x, y)
    return jsonify({"ok": True, "x": round(x, 2), "y": round(y, 2)})


@app.post("/api/gantry/step")
def api_gantry_set_step():
    """Set step size. Body: {step: float}"""
    data = request.get_json(force=True) or {}
    _gantry_state["step"] = max(0.1, min(50, float(data.get("step", 5.0))))
    return jsonify({"ok": True, "step": _gantry_state["step"]})


# ── camera ────────────────────────────────────────────────────────────────────

@app.post("/api/gantry/camera/open")
def api_gantry_camera_open():
    """Open the USB microscope camera."""
    if not HAS_CV2:
        return jsonify({"error": "opencv-python not installed"}), 501
    data = request.get_json(force=True) or {}
    idx = int(data.get("camera_index", GANTRY_DEFAULTS["camera_index"]))

    with _gantry_lock:
        if _gantry_state["camera"] is not None:
            _gantry_state["camera"].release()
        cap = cv2.VideoCapture(idx)
        if not cap.isOpened():
            return jsonify({"error": f"Could not open camera index {idx}"}), 400
        # Warmup
        for _ in range(GANTRY_DEFAULTS["warmup_frames"]):
            cap.read()
            time.sleep(0.1)
        _gantry_state["camera"] = cap
        _gantry_state["camera_ok"] = True
        w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    return jsonify({"ok": True, "width": w, "height": h})


@app.post("/api/gantry/camera/close")
def api_gantry_camera_close():
    """Close the camera."""
    with _gantry_lock:
        if _gantry_state["camera"] is not None:
            _gantry_state["camera"].release()
            _gantry_state["camera"] = None
            _gantry_state["camera_ok"] = False
    return jsonify({"ok": True})


@app.post("/api/gantry/camera/led")
def api_gantry_camera_led():
    """Toggle camera LED on/off (Dino-Lite ring light).

    Tries multiple approaches:
    1. OpenCV CAP_PROP_SETTINGS (opens native camera dialog on some drivers)
    2. Direct backlight property toggle
    3. Release & reopen camera as last resort to reset LED state
    """
    data = request.get_json(force=True) or {}
    turn_on = data.get("on")  # True/False, or None to toggle
    if turn_on is None:
        turn_on = not _gantry_state["led_on"]

    with _gantry_lock:
        cap = _gantry_state["camera"]
        if cap is None or not cap.isOpened():
            return jsonify({"error": "Camera not open. Open camera first."}), 400

        # Attempt 1: Use backlight property (works on some Dino-Lite models)
        # CAP_PROP_BACKLIGHT = 32
        try:
            cap.set(cv2.CAP_PROP_BACKLIGHT, 1.0 if turn_on else 0.0)
        except Exception:
            pass

        # Attempt 2: Use exposure to indirectly control — very low exposure
        # dims the image but doesn't turn off LEDs on most cameras.
        # Instead try auto-exposure off + manual exposure for "dark" effect.
        if not turn_on:
            try:
                # Disable auto-exposure (3=auto, 1=manual on many backends)
                cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 1)
                cap.set(cv2.CAP_PROP_EXPOSURE, -13)  # very dark
            except Exception:
                pass
        else:
            try:
                cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 3)  # re-enable auto
            except Exception:
                pass

        _gantry_state["led_on"] = turn_on

    return jsonify({"ok": True, "led_on": turn_on})


@app.get("/api/gantry/camera/snapshot")
def api_gantry_camera_snapshot():
    """Return a single JPEG snapshot from the camera."""
    if not _gantry_state["camera_ok"] or _gantry_state["camera"] is None:
        return jsonify({"error": "Camera not open"}), 400
    ret, frame = _gantry_state["camera"].read()
    if not ret or frame is None:
        return jsonify({"error": "Capture failed"}), 500
    _, buf = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
    return Response(buf.tobytes(), mimetype='image/jpeg')


@app.get("/api/gantry/camera/stream")
def api_gantry_camera_stream():
    """MJPEG stream from the USB microscope."""
    def generate():
        while _gantry_state["camera_ok"] and _gantry_state["camera"] is not None:
            ret, frame = _gantry_state["camera"].read()
            if not ret or frame is None:
                time.sleep(0.1)
                continue
            # Overlay position info
            x, y = _gantry_state["current_x"], _gantry_state["current_y"]
            cv2.putText(frame, f"X={x:.1f} Y={y:.1f}", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            _, buf = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')
            time.sleep(0.05)  # ~20 fps cap
    return Response(generate(), mimetype='multipart/x-mixed-replace; boundary=frame')


# ── positions ─────────────────────────────────────────────────────────────────

@app.post("/api/gantry/mark")
def api_gantry_mark():
    """Mark current position as a droplet. Optionally pass {id: "custom_name"}."""
    data = request.get_json(force=True) or {}
    count = len(_gantry_state["positions"]) + 1
    pos_id = data.get("id", f"droplet_{count:03d}")
    pos = {
        "id": pos_id,
        "x": round(_gantry_state["current_x"], 2),
        "y": round(_gantry_state["current_y"], 2),
    }
    _gantry_state["positions"].append(pos)

    # Save preview if camera available
    if _gantry_state["camera_ok"] and _gantry_state["camera"] is not None:
        ret, frame = _gantry_state["camera"].read()
        if ret and frame is not None:
            preview_dir = Path("calibration_previews")
            preview_dir.mkdir(parents=True, exist_ok=True)
            cv2.imwrite(str(preview_dir / f"{pos_id}.png"), frame)

    return jsonify({"ok": True, "position": pos, "total": len(_gantry_state["positions"])})


@app.delete("/api/gantry/positions/<pos_id>")
def api_gantry_remove_position(pos_id: str):
    """Remove a marked position by id."""
    before = len(_gantry_state["positions"])
    _gantry_state["positions"] = [p for p in _gantry_state["positions"] if p["id"] != pos_id]
    if len(_gantry_state["positions"]) == before:
        return jsonify({"error": "Position not found"}), 404
    return jsonify({"ok": True, "remaining": len(_gantry_state["positions"])})


@app.delete("/api/gantry/positions")
def api_gantry_clear_positions():
    """Remove all marked positions."""
    _gantry_state["positions"] = []
    return jsonify({"ok": True})


@app.post("/api/gantry/positions/save")
def api_gantry_save_positions():
    """Save positions to a JSON file. Body: {path: "positions.json"}"""
    data = request.get_json(force=True) or {}
    path = data.get("path", "positions.json")
    with open(path, 'w') as f:
        json.dump(_gantry_state["positions"], f, indent=2)
    return jsonify({"ok": True, "path": path, "count": len(_gantry_state["positions"])})


@app.post("/api/gantry/positions/load")
def api_gantry_load_positions():
    """Load positions from a JSON file. Body: {path: "positions.json"}"""
    data = request.get_json(force=True) or {}
    path = data.get("path", "positions.json")
    if not Path(path).exists():
        return jsonify({"error": f"File not found: {path}"}), 404
    with open(path) as f:
        _gantry_state["positions"] = json.load(f)
    return jsonify({"ok": True, "count": len(_gantry_state["positions"]),
                    "positions": _gantry_state["positions"]})


@app.post("/api/gantry/goto/<pos_id>")
def api_gantry_goto_position(pos_id: str):
    """Move gantry to a previously marked position."""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    pos = next((p for p in _gantry_state["positions"] if p["id"] == pos_id), None)
    if pos is None:
        return jsonify({"error": "Position not found"}), 404
    with _gantry_lock:
        _gantry_move_to(pos["x"], pos["y"])
    return jsonify({"ok": True, "x": pos["x"], "y": pos["y"]})


# ── scan & monitor ────────────────────────────────────────────────────────────

@app.post("/api/gantry/scan")
def api_gantry_scan():
    """One-shot scan: visit each position and capture an image.
    Body: {name: "exp_name"} — creates <name>_droplet_NNN experiments."""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    if not _gantry_state["camera_ok"]:
        return jsonify({"error": "Camera not open"}), 400
    if not _gantry_state["positions"]:
        return jsonify({"error": "No positions marked"}), 400

    data = request.get_json(force=True) or {}
    exp_name = data.get("name", f"scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    exp_base = Path(data.get("experiments_dir", str(DATA_ROOT)))

    captured = 0
    failed = 0

    for pos in _gantry_state["positions"]:
        with _gantry_lock:
            _gantry_move_to(pos["x"], pos["y"])
        time.sleep(GANTRY_DEFAULTS["settle_time"])

        ret, frame = _gantry_state["camera"].read()
        raw_dir = exp_base / exp_name / pos['id'] / "raw_images"
        raw_dir.mkdir(parents=True, exist_ok=True)

        if ret and frame is not None:
            fname = f"img_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            cv2.imwrite(str(raw_dir / fname), frame)
            captured += 1
        else:
            failed += 1

    return jsonify({
        "ok": True,
        "experiment": exp_name,
        "captured": captured,
        "failed": failed,
    })


def _monitor_loop(exp_name: str, exp_base: Path, interval: float,
                   duration_h: Optional[float]):
    """Background thread for continuous monitoring."""
    start_time = time.time()
    cycle = 0
    log = _gantry_state["monitor_log"]

    while _gantry_state["monitor_running"]:
        if duration_h:
            elapsed_h = (time.time() - start_time) / 3600
            if elapsed_h >= duration_h:
                log.append(f"Duration limit reached ({duration_h}h). Stopping.")
                break

        sweep_start = time.time()
        captured = 0

        for pos in _gantry_state["positions"]:
            if not _gantry_state["monitor_running"]:
                break
            with _gantry_lock:
                _gantry_move_to(pos["x"], pos["y"])
            time.sleep(GANTRY_DEFAULTS["settle_time"])

            if _gantry_state["camera"] is not None:
                ret, frame = _gantry_state["camera"].read()
                if ret and frame is not None:
                    raw_dir = exp_base / exp_name / pos['id'] / "raw_images"
                    raw_dir.mkdir(parents=True, exist_ok=True)
                    fname = f"img_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    cv2.imwrite(str(raw_dir / fname), frame)
                    captured += 1

        elapsed_total = (time.time() - start_time) / 3600
        msg = (f"Cycle {cycle}: {captured}/{len(_gantry_state['positions'])} captured, "
               f"elapsed {elapsed_total:.1f}h")
        log.append(msg)
        # Keep log bounded
        if len(log) > 200:
            del log[:100]
        cycle += 1

        # Wait for next sweep
        wait_time = max(0, interval - (time.time() - sweep_start))
        sleep_end = time.time() + wait_time
        while _gantry_state["monitor_running"] and time.time() < sleep_end:
            time.sleep(1)

    _gantry_state["monitor_running"] = False
    log.append("Monitor stopped.")


@app.post("/api/gantry/monitor/start")
def api_gantry_monitor_start():
    """Start continuous monitoring. Body: {name, interval, duration}"""
    if not _gantry_state["connected"]:
        return jsonify({"error": "Not connected"}), 400
    if not _gantry_state["camera_ok"]:
        return jsonify({"error": "Camera not open"}), 400
    if not _gantry_state["positions"]:
        return jsonify({"error": "No positions marked"}), 400
    if _gantry_state["monitor_running"]:
        return jsonify({"error": "Monitor already running"}), 400

    data = request.get_json(force=True) or {}
    exp_name = data.get("name", f"monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    interval = float(data.get("interval", 1800))
    duration = data.get("duration")
    if duration is not None:
        duration = float(duration)
    exp_base = Path(data.get("experiments_dir", str(DATA_ROOT)))

    # Create folders upfront
    for pos in _gantry_state["positions"]:
        (exp_base / exp_name / pos['id'] / "raw_images").mkdir(parents=True, exist_ok=True)

    _gantry_state["monitor_running"] = True
    _gantry_state["monitor_log"] = [f"Started: {exp_name}, interval={interval}s"]
    t = threading.Thread(
        target=_monitor_loop,
        args=(exp_name, exp_base, interval, duration),
        daemon=True,
    )
    _gantry_state["monitor_thread"] = t
    t.start()

    return jsonify({"ok": True, "experiment": exp_name})


@app.post("/api/gantry/monitor/stop")
def api_gantry_monitor_stop():
    """Stop the running monitor."""
    _gantry_state["monitor_running"] = False
    return jsonify({"ok": True})


# ─── main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n  Data root:  {DATA_ROOT}")
    print(f"  Experiments found: {list_experiments() or '(none yet)'}\n")
    print("  API Endpoints:")
    print("    GET  /api/experiments              - List all experiments")
    print("    GET  /api/frames/<exp_name>        - List frames for experiment")
    print("    GET  /api/crystals/<exp_name>      - List all crystals")
    print("    POST /api/crystals/<exp_name>      - Create a new crystal")
    print("    PUT  /api/crystals/<exp_name>/<id>/nucleation - Set nucleation time")
    print("    GET  /api/lengths/<exp_name>?crystal_id=<id> - Get measurements")
    print()
    print("  Pages:")
    print("    http://localhost:8000            - Crystal Annotator")
    print("    http://localhost:8000/gantry     - Gantry Controller\n")
    app.run(host="127.0.0.1", port=8000, debug=False)
